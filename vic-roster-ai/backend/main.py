from datetime import datetime
import os
import sqlite3
from typing import Dict, List, Tuple

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pulp import LpBinary, LpMinimize, LpProblem, LpVariable, lpSum, value
from pydantic import BaseModel

DB_FILENAME = "roster.db"
EXPORT_FILENAME = "Roster_Request.xlsx"

BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE_DIR, DB_FILENAME)
EXPORT_PATH = os.path.join(BASE_DIR, EXPORT_FILENAME)

DAYS = 14
SHIFTS = ["AM", "PM", "ND"]
SHIFT_CODE_MAP = {"AM": "D", "PM": "E", "ND": "N"}
SHIFT_BANNED_PAIRS = {("E", "D"), ("N", "D"), ("N", "E"), ("D", "N")}
ROLE_ORDER = ["ANUM", "CNS", "RN", "EN", "GNP"]
WEEKEND_INDEXES = {5, 6, 12, 13}
DEFAULT_REQUESTS = 2
DEFAULT_PREFERENCES = 2
BOOLEAN_FIELDS = [
    "flexibleWork",
    "swapWilling",
    "overtimeOptIn",
    "rightToDisconnectAck",
    "localInductionComplete",
]

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


class Profile(BaseModel):
    name: str
    email: str
    role: str = "RN"
    fte: str
    shiftPref: str
    maxNDs: str
    softLock: str = ""
    hardLock: str = ""
    cycle: str
    requestsQuota: int = DEFAULT_REQUESTS
    preferencesQuota: int = DEFAULT_PREFERENCES
    flexibleWork: bool = False
    swapWilling: bool = True
    overtimeOptIn: bool = False
    availabilityNotes: str = ""
    rightToDisconnectAck: bool = False
    localInductionComplete: bool = False
    supplementaryAvailability: str = ""


conn = sqlite3.connect(DB_PATH, check_same_thread=False)
conn.row_factory = sqlite3.Row

conn.execute(
    """
CREATE TABLE IF NOT EXISTS profiles (
    id INTEGER PRIMARY KEY,
    name TEXT,
    email TEXT UNIQUE,
    role TEXT DEFAULT 'RN',
    fte TEXT,
    shiftPref TEXT,
    maxNDs TEXT,
    softLock TEXT,
    hardLock TEXT,
    cycle TEXT,
    requests_quota INTEGER DEFAULT 2,
    preferences_quota INTEGER DEFAULT 2,
    flexible_work INTEGER DEFAULT 0,
    swap_willing INTEGER DEFAULT 1,
    overtime_opt_in INTEGER DEFAULT 0,
    availability_notes TEXT DEFAULT '',
    right_to_disconnect_ack INTEGER DEFAULT 0,
    local_induction_complete INTEGER DEFAULT 0,
    supplementary_availability TEXT DEFAULT '',
    submitted_at TEXT
)
"""
)

for column, dtype, default in [
    ("role", "TEXT", "'RN'"),
    ("requests_quota", "INTEGER", "2"),
    ("preferences_quota", "INTEGER", "2"),
    ("flexible_work", "INTEGER", "0"),
    ("swap_willing", "INTEGER", "1"),
    ("overtime_opt_in", "INTEGER", "0"),
    ("availability_notes", "TEXT", "''"),
    ("right_to_disconnect_ack", "INTEGER", "0"),
    ("local_induction_complete", "INTEGER", "0"),
    ("supplementary_availability", "TEXT", "''"),
]:
    try:
        conn.execute(f"ALTER TABLE profiles ADD COLUMN {column} {dtype} DEFAULT {default}")
    except sqlite3.OperationalError:
        pass
conn.commit()


def _bool_to_int(value: bool) -> int:
    return 1 if value else 0


def _int_to_bool(value) -> bool:
    return bool(value)


def validate_fte(value: str) -> Tuple[bool, str]:
    """Validate FTE is one of allowed values."""
    valid_ftes = ["0.6", "0.8", "1.0"]
    if value.strip() not in valid_ftes:
        return False, f"FTE must be one of {valid_ftes}, got '{value}'"
    return True, ""


def validate_role(value: str) -> Tuple[bool, str]:
    """Validate role is in ROLE_ORDER."""
    upper_role = value.upper()
    if upper_role not in ROLE_ORDER:
        return False, f"Role must be one of {ROLE_ORDER}, got '{value}'"
    return True, ""


def validate_max_nds(value: str) -> Tuple[bool, str]:
    """Validate maxNDs is 0-3 integer."""
    try:
        nd = int(value)
        if nd < 0 or nd > 3:
            return False, f"Max ND must be 0-3, got {nd}"
        return True, ""
    except ValueError:
        return False, f"Max ND must be integer, got '{value}'"


def validate_shift_pref(value: str) -> Tuple[bool, str]:
    """Validate shift preference is AM, PM, or ND."""
    if value not in SHIFTS:
        return False, f"Shift preference must be one of {SHIFTS}, got '{value}'"
    return True, ""


def validate_lock_format(value: str) -> Tuple[bool, str, int]:
    """Parse lock format 'DD MMM' and return (valid, error_msg, day_number)."""
    if not value or not value.strip():
        return True, "", -1  # Empty is OK
    
    try:
        # Expected format: "15 Nov" → extract day number
        parts = value.strip().split()
        if len(parts) < 1:
            return False, "Lock format must be 'DD MMM' (e.g., '15 Nov')", -1
        
        day = int(parts[0])
        if day < 1 or day > DAYS:
            return False, f"Lock day must be 1-{DAYS}, got {day}", -1
        
        return True, "", day
    except ValueError:
        return False, "Lock day must be numeric (e.g., '15 Nov')", -1


def validate_name(value: str) -> Tuple[bool, str]:
    """Validate name is not empty."""
    if not value or not value.strip():
        return False, "Name is required"
    return True, ""


def validate_email(value: str) -> Tuple[bool, str]:
    """Validate email format."""
    if not value or "@" not in value:
        return False, "Valid email is required"
    if len(value) < 5:
        return False, "Email too short"
    return True, ""


def fetch_profiles() -> List[Dict]:
    cur = conn.execute(
        """
        SELECT id, name, email, role, fte, shiftPref, maxNDs, softLock, hardLock, cycle,
               requests_quota, preferences_quota, flexible_work, swap_willing, overtime_opt_in,
               availability_notes, right_to_disconnect_ack, local_induction_complete,
               supplementary_availability, submitted_at
        FROM profiles
    """
    )
    profiles = []
    for row in cur.fetchall():
        profiles.append(
            {
                "id": row["id"],
                "name": row["name"],
                "email": row["email"],
                "role": (row["role"] or "RN").upper(),
                "fte": row["fte"],
                "shiftPref": row["shiftPref"],
                "maxNDs": row["maxNDs"],
                "softLock": row["softLock"] or "",
                "hardLock": row["hardLock"] or "",
                "cycle": row["cycle"],
                "requestsQuota": row["requests_quota"] or DEFAULT_REQUESTS,
                "preferencesQuota": row["preferences_quota"] or DEFAULT_PREFERENCES,
                "flexibleWork": _int_to_bool(row["flexible_work"]),
                "swapWilling": _int_to_bool(row["swap_willing"]),
                "overtimeOptIn": _int_to_bool(row["overtime_opt_in"]),
                "availabilityNotes": row["availability_notes"] or "",
                "rightToDisconnectAck": _int_to_bool(row["right_to_disconnect_ack"]),
                "localInductionComplete": _int_to_bool(row["local_induction_complete"]),
                "supplementaryAvailability": row["supplementary_availability"] or "",
                "submitted_at": row["submitted_at"],
            }
        )
    return profiles


def _role_sort_key(role: str) -> int:
    try:
        return ROLE_ORDER.index(role)
    except ValueError:
        return len(ROLE_ORDER)


def _build_roster_matrix(roster: List[Dict], names: List[str]) -> Dict[str, List[str]]:
    matrix = {name: ["OFF"] * DAYS for name in names}
    for day_index, day in enumerate(roster):
        for shift in SHIFTS:
            code = SHIFT_CODE_MAP[shift]
            for name in day[shift]:
                matrix.setdefault(name, ["OFF"] * DAYS)
                matrix[name][day_index] = code
    return matrix


def _compute_analytics(matrix: Dict[str, List[str]], profiles: Dict[str, Dict]) -> Tuple[List[Dict], Dict]:
    analytics = []
    warnings = []
    overall_ok = True

    for name, shifts in matrix.items():
        profile = profiles.get(name)
        if not profile:
            continue
        fte = float(profile["fte"])
        weekend_count = sum(1 for idx, shift in enumerate(shifts) if idx in WEEKEND_INDEXES and shift != "OFF")

        max_consecutive = 0
        current_consecutive = 0
        longest_off = 0
        current_off = 0
        rest_breaches = []

        for idx, shift in enumerate(shifts):
            if shift == "OFF":
                current_consecutive = 0
                current_off += 1
                longest_off = max(longest_off, current_off)
            else:
                current_consecutive += 1
                max_consecutive = max(max_consecutive, current_consecutive)
                current_off = 0
                if idx < DAYS - 1:
                    nxt = shifts[idx + 1]
                    if nxt != "OFF" and (shift, nxt) in SHIFT_BANNED_PAIRS:
                        rest_breaches.append((idx, f"{shift}->{nxt}"))

        two_day_break = longest_off >= 2
        consecutive_ok = max_consecutive <= 6
        rest_ok = len(rest_breaches) == 0
        fatigue_score = 0
        notes = []

        if not consecutive_ok:
            fatigue_score += 2
            notes.append("More than six consecutive shifts")
        if not rest_ok:
            fatigue_score += len(rest_breaches)
            notes.append("Turnaround breach (<10h) detected")
        if weekend_count > 4 and not profile["flexibleWork"]:
            fatigue_score += weekend_count - 4
            notes.append("High weekend workload")
        if fte >= 0.8 and not two_day_break and not profile["flexibleWork"]:
            fatigue_score += 2
            notes.append("Missing two consecutive days off")

        compliant = consecutive_ok and rest_ok and (fte < 0.8 or two_day_break or profile["flexibleWork"])

        if not compliant:
            overall_ok = False
            warnings.append({"name": name, "issues": notes})

        analytics.append(
            {
                "name": name,
                "role": profile["role"],
                "shiftPref": profile["shiftPref"],
                "fte": fte,
                "weekendCount": weekend_count,
                "maxConsecutive": max_consecutive,
                "longestOffStreak": longest_off,
                "hasTwoDayBreak": two_day_break,
                "restBreaches": rest_breaches,
                "fatigueScore": fatigue_score,
                "flexibleWork": profile["flexibleWork"],
                "swapWilling": profile["swapWilling"],
                "overtimeOptIn": profile["overtimeOptIn"],
                "compliant": compliant,
                "notes": notes,
            }
        )

    analytics.sort(key=lambda item: (_role_sort_key(item["role"]), item["name"]))

    return analytics, {
        "overall": "pass" if overall_ok else "attention",
        "warnings": warnings,
    }


@app.post("/submit-profile")
def submit_profile(profile: Profile):
    # Validation step 1: Right to Disconnect acknowledgement
    if not profile.rightToDisconnectAck:
        raise HTTPException(status_code=400, detail="Right to Disconnect acknowledgement is required.")
    
    # Validation step 2: Name
    valid, error_msg = validate_name(profile.name)
    if not valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Validation step 3: Email
    valid, error_msg = validate_email(profile.email)
    if not valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Validation step 4: Role
    valid, error_msg = validate_role(profile.role)
    if not valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Validation step 5: FTE
    valid, error_msg = validate_fte(profile.fte)
    if not valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Validation step 6: Shift preference
    valid, error_msg = validate_shift_pref(profile.shiftPref)
    if not valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Validation step 7: Max ND
    valid, error_msg = validate_max_nds(profile.maxNDs)
    if not valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Validation step 8: Soft lock format
    valid, error_msg, _ = validate_lock_format(profile.softLock)
    if not valid:
        raise HTTPException(status_code=400, detail=f"Soft lock: {error_msg}")
    
    # Validation step 9: Hard lock format
    valid, error_msg, _ = validate_lock_format(profile.hardLock)
    if not valid:
        raise HTTPException(status_code=400, detail=f"Hard lock: {error_msg}")
    
    # Validation step 10: Request/preference quotas
    if profile.requestsQuota < 0 or profile.requestsQuota > 4:
        raise HTTPException(status_code=400, detail="Requests quota must be 0-4")
    if profile.preferencesQuota < 0 or profile.preferencesQuota > 4:
        raise HTTPException(status_code=400, detail="Preferences quota must be 0-4")

    try:
        conn.execute(
            """
            INSERT INTO profiles (
                name, email, role, fte, shiftPref, maxNDs, softLock, hardLock, cycle,
                requests_quota, preferences_quota, flexible_work, swap_willing, overtime_opt_in,
                availability_notes, right_to_disconnect_ack, local_induction_complete,
                supplementary_availability, submitted_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                profile.name.strip(),
                profile.email.strip(),
                profile.role.upper(),
                profile.fte,
                profile.shiftPref,
                profile.maxNDs,
                profile.softLock,
                profile.hardLock,
                profile.cycle,
                profile.requestsQuota,
                profile.preferencesQuota,
                _bool_to_int(profile.flexibleWork),
                _bool_to_int(profile.swapWilling),
                _bool_to_int(profile.overtimeOptIn),
                profile.availabilityNotes,
                _bool_to_int(profile.rightToDisconnectAck),
                _bool_to_int(profile.localInductionComplete),
                profile.supplementaryAvailability,
                datetime.now().isoformat(),
            ),
        )
        conn.commit()
        return {"status": "success"}
    except sqlite3.IntegrityError:
        raise HTTPException(400, "Email already submitted")
    except Exception as e:
        raise HTTPException(500, f"Database error: {str(e)}")


@app.get("/profiles")
def get_profiles():
    return fetch_profiles()


@app.get("/generate-roster")
def generate_roster():
    profiles = fetch_profiles()
    if not profiles:
        raise HTTPException(400, "No profiles")

    staff_names = [p["name"] for p in profiles]
    name_index = {name: idx for idx, name in enumerate(staff_names)}

    prob = LpProblem("Roster", LpMinimize)

    x = LpVariable.dicts("assign", (range(len(staff_names)), range(DAYS), range(len(SHIFTS))), cat=LpBinary)

    for i in range(len(staff_names)):
        for d in range(DAYS):
            prob += lpSum(x[i][d][k] for k in range(len(SHIFTS))) <= 1

    for d in range(DAYS):
        for k in range(len(SHIFTS)):
            prob += lpSum(x[i][d][k] for i in range(len(staff_names))) >= 1

    for i, profile in enumerate(profiles):
        fte = float(profile["fte"])
        target = int(round(fte * DAYS))
        work_total = lpSum(x[i][d][k] for d in range(DAYS) for k in range(len(SHIFTS)))
        prob += work_total >= max(target - 1, 0)
        prob += work_total <= target + 1

    for i, profile in enumerate(profiles):
        for start in range(DAYS - 6):
            prob += lpSum(x[i][d][k] for d in range(start, start + 7) for k in range(len(SHIFTS))) <= 6

        max_nds = int(profile["maxNDs"])
        prob += lpSum(x[i][d][2] for d in range(DAYS)) <= max_nds

        soft_lock = profile["softLock"]
        hard_lock = profile["hardLock"]

        if hard_lock:
            try:
                day = int(hard_lock.split("-")[-1]) - 1
                if 0 <= day < DAYS:
                    prob += lpSum(x[i][day][k] for k in range(len(SHIFTS))) == 0
            except ValueError:
                pass

        if soft_lock:
            try:
                day = int(soft_lock.split("-")[-1]) - 1
                if 0 <= day < DAYS:
                    prob += lpSum(x[i][day][k] for k in range(len(SHIFTS))) <= 0
            except ValueError:
                pass

        for d in range(DAYS - 1):
            for k1, shift1 in enumerate(SHIFTS):
                for k2, shift2 in enumerate(SHIFTS):
                    if (SHIFT_CODE_MAP[shift1], SHIFT_CODE_MAP[shift2]) in SHIFT_BANNED_PAIRS:
                        prob += x[i][d][k1] + x[i][d + 1][k2] <= 1

    pref_penalty = lpSum(
        x[name_index[profile["name"]]][d][k]
        for profile in profiles
        for d in range(DAYS)
        for k in range(len(SHIFTS))
        if SHIFTS[k] != profile["shiftPref"]
    )
    prob += pref_penalty

    prob.solve()
    if prob.status != 1:
        return {"status": "infeasible", "message": "No feasible roster with current constraints"}

    roster: List[Dict] = []
    for d in range(DAYS):
        day_entry = {"day": d + 1, "AM": [], "PM": [], "ND": []}
        for i, profile in enumerate(profiles):
            for k, shift in enumerate(SHIFTS):
                if value(x[i][d][k]) == 1:
                    day_entry[shift].append(profile["name"])
        roster.append(day_entry)

    matrix = _build_roster_matrix(roster, staff_names)
    profile_map = {p["name"]: p for p in profiles}
    analytics, compliance = _compute_analytics(matrix, profile_map)

    return {
        "status": "valid",
        "roster": roster,
        "analytics": analytics,
        "compliance": compliance,
    }


@app.get("/export-excel")
def export_excel():
    data = generate_roster()
    if not isinstance(data, dict):
        return data
    if data.get("status") != "valid":
        raise HTTPException(400, data.get("message", "Roster not compliant"))

    roster = data["roster"]
    analytics = data.get("analytics", [])
    profiles = fetch_profiles()
    profile_map = {p["name"]: p for p in profiles}

    staff_names = sorted(profile_map.keys(), key=lambda name: (_role_sort_key(profile_map[name]["role"]), name))
    matrix = _build_roster_matrix(roster, staff_names)

    shift_fill = {
        "D": PatternFill("solid", fgColor="FFE4B5"),    # Day: Moccasin (warm orange)
        "E": PatternFill("solid", fgColor="FFED9E"),    # Evening: Light yellow
        "N": PatternFill("solid", fgColor="2E7D32"),    # Night: Deep green
        "OFF": PatternFill("solid", fgColor="E9ECEF"),  # Off: Light gray
    }
    # Font colors for better contrast
    shift_font = {
        "D": Font(bold=True, color="1F1F1F"),           # Day: Dark text
        "E": Font(bold=True, color="333333"),           # Evening: Dark gray text
        "N": Font(bold=True, color="FFFFFF"),           # Night: White text
        "OFF": Font(color="666666"),                    # Off: Gray text
    }
    thin = Side(border_style="thin", color="D0D0D0")
    thick = Side(border_style="medium", color="000000")
    header_band = PatternFill("solid", fgColor="BBD0F7")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Published Roster"

    sheet.merge_cells("A1:P1")
    sheet["A1"] = "Published Roster – Ward A – Fortnight"
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"].fill = PatternFill("solid", fgColor="4B0082")

    sheet.merge_cells("A2:M2")
    sheet["A2"] = "Shift Codes: D = Day (0700–1530), E = Evening (1300–2130), N = Night (2100–0730), OFF = Day Off"
    sheet["A2"].alignment = Alignment(horizontal="left")
    sheet["A2"].font = Font(size=11)

    sheet.merge_cells("N2:P2")
    sheet["N2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    sheet["N2"].alignment = Alignment(horizontal="center")
    sheet["N2"].font = Font(bold=True, color="FFFFFF")
    sheet["N2"].fill = PatternFill("solid", fgColor="6A1B9A")

    sheet.merge_cells("A3:P3")
    sheet["A3"] = "All requests resolved. Retain for seven (7) years."
    sheet["A3"].alignment = Alignment(horizontal="center")
    sheet["A3"].font = Font(italic=True, size=11)

    sheet["A5"] = "Role"
    sheet["B5"] = "Name"
    sheet["C5"] = "Compliance"
    for label in ("A5", "B5", "C5"):
        cell = sheet[label]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_band

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 26
    sheet.column_dimensions["C"].width = 32

    sheet.merge_cells("D4:J4")
    sheet["D4"] = "WEEK 1"
    sheet["D4"].font = Font(bold=True, color="FFFFFF")
    sheet["D4"].alignment = Alignment(horizontal="center")
    sheet["D4"].fill = PatternFill("solid", fgColor="6C757D")

    sheet.merge_cells("K4:Q4")
    sheet["K4"] = "WEEK 2"
    sheet["K4"].font = Font(bold=True, color="FFFFFF")
    sheet["K4"].alignment = Alignment(horizontal="center")
    sheet["K4"].fill = PatternFill("solid", fgColor="6C757D")

    days_labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for idx, label in enumerate(days_labels):
        col = chr(ord("D") + idx)
        sheet[f"{col}5"] = label
        sheet[f"{col}5"].font = Font(bold=True)
        sheet[f"{col}5"].alignment = Alignment(horizontal="center")
        sheet[f"{col}5"].fill = header_band
        sheet.column_dimensions[col].width = 11

        col_week2 = chr(ord("K") + idx)
        sheet[f"{col_week2}5"] = label
        sheet[f"{col_week2}5"].font = Font(bold=True)
        sheet[f"{col_week2}5"].alignment = Alignment(horizontal="center")
        sheet[f"{col_week2}5"].fill = header_band
        sheet.column_dimensions[col_week2].width = 11

    start_row = 6
    for row_idx, name in enumerate(staff_names, start=start_row):
        profile = profile_map[name]
        row = matrix[name]
        analytics_entry = next((a for a in analytics if a["name"] == name), None)

        sheet[f"A{row_idx}"] = profile["role"]
        sheet[f"A{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
        sheet[f"A{row_idx}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        sheet[f"B{row_idx}"] = name
        sheet[f"B{row_idx}"].alignment = Alignment(vertical="center")
        sheet[f"B{row_idx}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        compliance_text = "Compliant"
        if analytics_entry and not analytics_entry["compliant"]:
            compliance_text = "; ".join(analytics_entry["notes"])
        sheet[f"C{row_idx}"] = compliance_text or "Compliant"
        sheet[f"C{row_idx}"].alignment = Alignment(vertical="center")
        sheet[f"C{row_idx}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for day_idx, shift in enumerate(row):
            col_letter = chr(ord("D") + day_idx) if day_idx < 7 else chr(ord("K") + (day_idx - 7))
            cell = sheet[f"{col_letter}{row_idx}"]
            cell.value = shift
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            fill = shift_fill.get(shift, shift_fill["OFF"])
            cell.fill = fill
            cell.font = shift_font.get(shift, shift_font["OFF"])  # Apply appropriate font for shift type

    sheet_compliance = workbook.create_sheet("Compliance Summary")
    sheet_compliance.append(
        [
            "Name",
            "Role",
            "FTE",
            "Weekend Shifts",
            "Max Consecutive",
            "Longest Off Streak",
            "Two-Day Break",
            "Rest Breaches",
            "Fatigue Score",
            "Notes",
        ]
    )

    for cell in sheet_compliance[1]:
        cell.font = Font(bold=True)
        cell.fill = header_band
        cell.alignment = Alignment(horizontal="center")

    for entry in analytics:
        sheet_compliance.append(
            [
                entry["name"],
                entry["role"],
                entry["fte"],
                entry["weekendCount"],
                entry["maxConsecutive"],
                entry["longestOffStreak"],
                "Yes" if entry["hasTwoDayBreak"] else "No",
                ", ".join(f"{idx + 1}:{label}" for idx, label in entry["restBreaches"]) or "-",
                entry["fatigueScore"],
                "; ".join(entry["notes"]) or "-",
            ]
        )

    # Add footer with metadata
    footer_row = max([row for row, name in enumerate(staff_names, start=6)]) + 2
    sheet.merge_cells(f"A{footer_row}:Q{footer_row}")
    footer_cell = sheet[f"A{footer_row}"]
    footer_cell.value = (
        f"Document ID: GDL-25994 | Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
        f"Compliant rosters: {sum(1 for a in analytics if a.get('compliant', True))}/{len(analytics)}"
    )
    footer_cell.alignment = Alignment(horizontal="center")
    footer_cell.font = Font(size=9, color="666666", italic=True)

    workbook.save(EXPORT_PATH)
    return FileResponse(
        EXPORT_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=EXPORT_FILENAME,
    )
